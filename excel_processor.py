import pandas as pd
import numpy as np
from datetime import datetime, date
import os
import tempfile
import re

class ExcelProcessor:
    def __init__(self):
        # Definisi kolom output sesuai format yang diminta
        self.output_columns = [
            'PROVID',
            'PROVIDER_NAME',
            'SERVICECODE',
            'SERVICECODE DESCRIPTION',
            'KELAS',
            'RUANG BEDAH (SURGERY)/NON RUANG BEDAH (NON SURGERY)',
            'HELPER',
            'TARIFF',
            'TARIFF DESCRIPTION',
            'QUANTITY',
            'TOTAL BILLED',
            'GIVEN DATE (month, day, year)',
            'HEAMODIALISA/CHEMOTHERAPY/ODC/PHYSIOTHERAPY/RADIOTHERAPY',
            'HEAMODIALISA/CHEMOTHERAPY/ODC/PHYSIOTHERAPY/RADIOTHERAPY DESCRIPTION',
            'ICD_X_DIAGNOSIS_PRIMARY',
            'ICD_X_DESC_PRIMARY',
            'ICD_X_DIAGNOSIS_SECONDARY',
            'ICD_X_DESC_SECONDARY',
            'PHYSICIAN NAME',
            'PHYSICIAN DESCRIPTION (DPJP/IGD/POLICLINIC)',
            'CLIENT NAME',
            'CLIENTS DOB (month, day, year)',
            'CLIENTS SEX',
            'CLIENTS ADDRESS',
            'CLIENTS MEMBER ID',
            'CLIENTS MR NUMBER',
            'CLIENTS INVOICE NUMBER',
            'CLIENTSREGISTER NUMBER',
            'CLIENTS OTHER NUMBER',
            'admission',
            'discharge',
            'LoS'
        ]
        
        # Mapping untuk field yang ditemukan dalam data
        self.field_mapping = {
            'nomor_tagihan': ['nomor tagihan', 'no tagihan', 'invoice number', 'bill number'],
            'nomor_registrasi': ['nomor registrasi', 'no registrasi', 'registration number', 'reg number'],
            'tanggal_registrasi': ['tanggal registrasi', 'tgl registrasi', 'registration date', 'reg date'],
            'penjamin_bayar': ['penjamin bayar', 'asuransi', 'insurance', 'guarantor'],
            'nama_pasien': ['nama pasien', 'nama', 'patient name', 'pasien'],
            'terima_dari': ['terima dari', 'dari', 'from', 'received from'],
            'kelas_kamar': ['kelas kamar', 'kelas', 'kamar', 'room class', 'class'],
            'tanggal_keluar': ['tanggal keluar', 'tgl keluar', 'discharge date', 'exit date'],
            'kelas_dijamin': ['kelas dijamin', 'kelas asuransi', 'insured class'],
            'jenis_biaya': ['jenis biaya', 'jenis', 'biaya', 'cost type', 'expense type'],
            'waktu': ['waktu', 'time', 'jam', 'hour'],
            'tanggal': ['tanggal', 'tgl', 'date'],
            'keterangan': ['keterangan', 'deskripsi', 'description', 'note'],
            'jumlah': ['jumlah', 'qty', 'quantity', 'qty'],
            'nilai': ['nilai', 'harga', 'price', 'amount', 'tarif'],
            'sub_total': ['sub total', 'subtotal', 'total', 'sum']
        }
    
    def preview_excel(self, filepath):
        """Membaca dan menganalisis struktur data Excel secara mendalam"""
        try:
            # Baca file Excel dengan berbagai sheet
            excel_file = pd.ExcelFile(filepath)
            sheet_names = excel_file.sheet_names
            
            print(f"📊 Menganalisis {len(sheet_names)} sheet: {sheet_names}")
            
            if not sheet_names:
                raise Exception("File Excel tidak memiliki sheet")
            
            # Analisis mendalam untuk setiap sheet
            all_analysis = {}
            for sheet_name in sheet_names:
                try:
                    print(f"🔍 Menganalisis sheet: {sheet_name}")
                    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
                    
                    if df.empty:
                        print(f"⚠️ Warning: Sheet '{sheet_name}' kosong")
                        continue
                    
                    analysis = self._deep_analyze_sheet(df, sheet_name)
                    all_analysis[sheet_name] = analysis
                    
                except Exception as sheet_error:
                    print(f"⚠️ Warning: Error analyzing sheet '{sheet_name}': {sheet_error}")
                    # Continue with other sheets instead of failing completely
                    continue
            
            if not all_analysis:
                raise Exception("Tidak ada sheet yang dapat dianalisis")
            
            # Gabungkan analisis dari semua sheet
            combined_analysis = self._combine_sheet_analysis(all_analysis)
            
            # Bersihkan data untuk JSON serialization
            cleaned_analysis = self._clean_analysis_for_json(combined_analysis)
            
            return cleaned_analysis
            
        except Exception as e:
            raise Exception(f"Error membaca file Excel: {str(e)}")
    
    def _deep_analyze_sheet(self, df, sheet_name):
        """Analisis mendalam untuk satu sheet"""
        try:
            analysis = {
                'sheet_name': sheet_name,
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'detected_fields': {},
                'data_patterns': {},
                'sample_data': {},
                'header_rows': [],
                'data_rows': []
            }
            
            # Safety check for empty dataframe
            if df.empty or len(df.columns) == 0:
                print(f"⚠️ Warning: Sheet '{sheet_name}' is empty or has no columns")
                return analysis
            
            # Deteksi header rows (baris yang berisi label field)
            header_rows = self._detect_header_rows(df)
            analysis['header_rows'] = header_rows
            
            # Deteksi data rows (baris yang berisi nilai)
            data_rows = self._detect_data_rows(df, header_rows)
            analysis['data_rows'] = data_rows
            
            # Analisis setiap kolom untuk pola data
            for col_idx in range(len(df.columns)):
                try:
                    col_data = df.iloc[:, col_idx].dropna()
                    if len(col_data) > 0:
                        col_analysis = self._analyze_column_pattern(col_data, col_idx)
                        analysis['data_patterns'][f'col_{col_idx}'] = col_analysis
                        
                        # Deteksi field berdasarkan konten
                        detected_field = self._detect_field_from_content(col_data)
                        if detected_field:
                            analysis['detected_fields'][f'col_{col_idx}'] = detected_field
                        
                        # Sample data
                        analysis['sample_data'][f'col_{col_idx}'] = col_data.head(5).tolist()
                except Exception as col_error:
                    print(f"⚠️ Warning: Error analyzing column {col_idx}: {col_error}")
                    continue
            
            return analysis
            
        except Exception as e:
            print(f"⚠️ Warning: Error in deep analysis of sheet '{sheet_name}': {e}")
            # Return minimal analysis structure
            return {
                'sheet_name': sheet_name,
                'total_rows': len(df) if not df.empty else 0,
                'total_columns': len(df.columns) if not df.empty else 0,
                'detected_fields': {},
                'data_patterns': {},
                'sample_data': {},
                'header_rows': [],
                'data_rows': []
            }
    
    def _detect_header_rows(self, df):
        """Deteksi baris yang berisi header/label field"""
        header_rows = []
        
        for row_idx in range(min(10, len(df))):  # Cek 10 baris pertama
            row_data = df.iloc[row_idx].astype(str)
            header_score = self._calculate_header_score(row_data)
            
            if header_score > 0.6:  # Threshold untuk mendeteksi header
                header_rows.append({
                    'row_index': row_idx,
                    'score': header_score,
                    'content': row_data.tolist()
                })
        
        return sorted(header_rows, key=lambda x: x['score'], reverse=True)
    
    def _calculate_header_score(self, row_data):
        """Hitung skor kemungkinan baris adalah header"""
        score = 0
        total_cells = len(row_data)
        
        for cell_value in row_data:
            cell_str = str(cell_value).lower().strip()
            
            # Cek apakah cell berisi label field yang umum
            if any(field in cell_str for field_list in self.field_mapping.values() for field in field_list):
                score += 1
            elif any(keyword in cell_str for keyword in ['nomor', 'tanggal', 'nama', 'kelas', 'biaya', 'jumlah', 'total']):
                score += 0.5
            elif len(cell_str) > 3 and len(cell_str) < 50:  # Panjang yang masuk akal untuk header
                score += 0.3
        
        return score / total_cells if total_cells > 0 else 0
    
    def _detect_data_rows(self, df, header_rows):
        """Deteksi baris yang berisi data berdasarkan header rows"""
        data_rows = []
        
        if not header_rows:
            # Jika tidak ada header yang jelas, asumsikan semua baris adalah data
            return [{'row_index': i, 'type': 'data'} for i in range(len(df))]
        
        # Ambil header row pertama sebagai referensi
        header_row_idx = header_rows[0]['row_index']
        
        for row_idx in range(len(df)):
            if row_idx != header_row_idx:
                row_data = df.iloc[row_idx]
                if self._is_data_row(row_data):
                    data_rows.append({
                        'row_index': row_idx,
                        'type': 'data',
                        'content': row_data.tolist()
                    })
        
        return data_rows
    
    def _is_data_row(self, row_data):
        """Cek apakah baris berisi data (bukan header)"""
        # Cek apakah ada nilai numerik atau tanggal
        numeric_count = 0
        date_count = 0
        
        for cell_value in row_data:
            if pd.notna(cell_value):
                if isinstance(cell_value, (int, float)):
                    numeric_count += 1
                elif isinstance(cell_value, str):
                    # Cek apakah string berisi angka
                    if re.search(r'\d', str(cell_value)):
                        numeric_count += 1
                    # Cek apakah string berisi format tanggal
                    if re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', str(cell_value)):
                        date_count += 1
        
        # Baris dianggap data jika ada cukup nilai numerik atau tanggal
        return numeric_count > 0 or date_count > 0
    
    def _analyze_column_pattern(self, col_data, col_idx):
        """Analisis pola data dalam satu kolom"""
        try:
            pattern = {
                'data_type': str(col_data.dtype),
                'non_null_count': len(col_data),
                'unique_values': col_data.nunique(),
                'pattern_type': 'unknown'
            }
            
            # Safety check for empty column
            if col_data.empty:
                return pattern
            
            # Deteksi tipe pola
            try:
                if col_data.dtype in ['int64', 'float64']:
                    pattern['pattern_type'] = 'numeric'
                elif hasattr(col_data, 'str') and col_data.str.contains(r'Rp\s*[\d,]+', na=False).any():
                    pattern['pattern_type'] = 'currency'
                elif hasattr(col_data, 'str') and col_data.str.contains(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', na=False).any():
                    pattern['pattern_type'] = 'date'
                elif hasattr(col_data, 'str') and col_data.str.contains(r'^\d+$', na=False).any():
                    pattern['pattern_type'] = 'numeric_string'
                else:
                    pattern['pattern_type'] = 'text'
            except Exception as pattern_error:
                print(f"⚠️ Warning: Error detecting pattern for column {col_idx}: {pattern_error}")
                pattern['pattern_type'] = 'unknown'
            
            return pattern
            
        except Exception as e:
            print(f"⚠️ Warning: Error analyzing column pattern for column {col_idx}: {e}")
            return {
                'data_type': 'unknown',
                'non_null_count': 0,
                'unique_values': 0,
                'pattern_type': 'unknown'
            }
    
    def _detect_field_from_content(self, col_data):
        """Deteksi field berdasarkan konten kolom"""
        try:
            # Safety check for empty column
            if col_data.empty:
                return None
            
            # Gabungkan semua nilai dalam kolom
            combined_text = ' '.join(col_data.astype(str)).lower()
            
            # Cek setiap field mapping
            for field_name, keywords in self.field_mapping.items():
                if any(keyword in combined_text for keyword in keywords):
                    return field_name
            
            return None
            
        except Exception as e:
            print(f"⚠️ Warning: Error detecting field from content: {e}")
            return None
    
    def _combine_sheet_analysis(self, all_analysis):
        """Gabungkan analisis dari semua sheet"""
        try:
            combined = {
                'total_sheets': len(all_analysis),
                'sheets': all_analysis,
                'global_detected_fields': {},
                'summary': {}
            }
            
            # Gabungkan field yang terdeteksi dari semua sheet
            all_fields = {}
            for sheet_name, analysis in all_analysis.items():
                try:
                    for col_name, field in analysis.get('detected_fields', {}).items():
                        if field not in all_fields:
                            all_fields[field] = []
                        all_fields[field].append({
                            'sheet': sheet_name,
                            'column': col_name
                        })
                except Exception as sheet_error:
                    print(f"⚠️ Warning: Error processing sheet '{sheet_name}' in combine: {sheet_error}")
                    continue
            
            combined['global_detected_fields'] = all_fields
            
            # Buat summary dengan safety checks
            try:
                total_rows = sum(analysis.get('total_rows', 0) for analysis in all_analysis.values())
                total_columns = max((analysis.get('total_columns', 0) for analysis in all_analysis.values()), default=0)
                
                combined['summary'] = {
                    'total_rows': total_rows,
                    'total_columns': total_columns,
                    'detected_field_count': len(all_fields)
                }
            except Exception as summary_error:
                print(f"⚠️ Warning: Error creating summary: {summary_error}")
                combined['summary'] = {
                    'total_rows': 0,
                    'total_columns': 0,
                    'detected_field_count': len(all_fields)
                }
            
            return combined
            
        except Exception as e:
            print(f"⚠️ Warning: Error in combine sheet analysis: {e}")
            # Return minimal structure
            return {
                'total_sheets': len(all_analysis),
                'sheets': all_analysis,
                'global_detected_fields': {},
                'summary': {
                    'total_rows': 0,
                    'total_columns': 0,
                    'detected_field_count': 0
                }
            }
    
    def _clean_analysis_for_json(self, analysis):
        """Bersihkan analisis untuk JSON serialization"""
        try:
            def clean_value(value):
                try:
                    if pd.isna(value):
                        return ''
                    elif isinstance(value, (int, float)):
                        if np.isinf(value):
                            return ''
                        return str(value)
                    else:
                        return str(value).replace('\x00', '').replace('\n', ' ').replace('\r', '')
                except Exception:
                    return ''
            
            # Bersihkan sample data
            if 'sheets' in analysis:
                for sheet_name, sheet_analysis in analysis['sheets'].items():
                    try:
                        if 'sample_data' in sheet_analysis:
                            for col_name, sample_data in sheet_analysis['sample_data'].items():
                                try:
                                    sheet_analysis['sample_data'][col_name] = [clean_value(val) for val in sample_data]
                                except Exception as col_error:
                                    print(f"⚠️ Warning: Error cleaning column {col_name}: {col_error}")
                                    sheet_analysis['sample_data'][col_name] = []
                    except Exception as sheet_error:
                        print(f"⚠️ Warning: Error cleaning sheet {sheet_name}: {sheet_error}")
                        continue
            
            return analysis
            
        except Exception as e:
            print(f"⚠️ Warning: Error in JSON cleaning: {e}")
            return analysis
    
    def process_excel(self, filepath, options=None):
        """Memproses file Excel dengan analisis mendalam"""
        try:
            # Analisis mendalam terlebih dahulu
            analysis = self.preview_excel(filepath)
            
            # Baca data berdasarkan analisis
            processed_data = self._extract_structured_data(filepath, analysis)
            
            # Transform ke format output
            output_df = self._transform_to_output_format(processed_data, analysis)
            
            # Buat file output
            output_filepath = self._create_output_file(output_df, filepath)
            
            return output_filepath
            
        except Exception as e:
            raise Exception(f"Error memproses file Excel: {str(e)}")
    
    def _extract_structured_data(self, filepath, analysis):
        """Ekstrak data terstruktur berdasarkan analisis"""
        extracted_data = []
        
        for sheet_name, sheet_analysis in analysis['sheets'].items():
            print(f"📊 Memproses sheet: {sheet_name}")
            
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
            
            # Cek apakah ini format key-value pairs atau format tabel standar
            if self._is_key_value_format(df):
                print(f"🔍 Detected key-value format in sheet: {sheet_name}")
                sheet_data = self._extract_key_value_data(df, sheet_name)
            else:
                print(f"🔍 Detected standard table format in sheet: {sheet_name}")
                # Gunakan header rows yang terdeteksi
                if sheet_analysis['header_rows']:
                    header_row = sheet_analysis['header_rows'][0]
                    header_idx = header_row['row_index']
                    
                    # Baca data dengan header yang benar
                    data_df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_idx)
                    
                    # Bersihkan nama kolom
                    data_df.columns = [str(col).strip() for col in data_df.columns]
                    
                    # Ekstrak data berdasarkan field yang terdeteksi
                    sheet_data = self._extract_sheet_data(data_df, sheet_analysis)
                else:
                    print(f"⚠️ No header rows detected, using raw data")
                    sheet_data = self._extract_raw_data(df, sheet_name)
            
            extracted_data.extend(sheet_data)
        
        return extracted_data
    
    def _extract_sheet_data(self, df, sheet_analysis):
        """Ekstrak data dari satu sheet"""
        extracted_rows = []
        
        # Mapping kolom berdasarkan field yang terdeteksi
        field_columns = {}
        for col_name, field in sheet_analysis.get('detected_fields', {}).items():
            col_idx = int(col_name.split('_')[1])
            if col_idx < len(df.columns):
                field_columns[field] = df.columns[col_idx]
        
        # Proses setiap baris data
        for idx, row in df.iterrows():
            if idx == 0:  # Skip header row
                continue
                
            extracted_row = {}
            
            # Ekstrak data berdasarkan field mapping
            for field, col_name in field_columns.items():
                if col_name in df.columns:
                    extracted_row[field] = self._clean_value(row[col_name])
            
            if extracted_row:  # Hanya tambahkan jika ada data
                extracted_rows.append(extracted_row)
        
        return extracted_rows
    
    def _is_key_value_format(self, df):
        """Deteksi apakah data dalam format key-value pairs"""
        try:
            # Cek apakah ada pola key-value seperti "Nomor Tagihan : IP-00030178"
            key_value_pattern = False
            for col in df.columns:
                col_data = df.iloc[:, col].dropna()
                if len(col_data) > 0:
                    # Cek apakah ada string yang mengandung " : " (key-value separator)
                    if any(' : ' in str(val) for val in col_data if pd.notna(val)):
                        key_value_pattern = True
                        break
            
            return key_value_pattern
        except Exception as e:
            print(f"⚠️ Warning: Error detecting key-value format: {e}")
            return False
    
    def _extract_key_value_data(self, df, sheet_name):
        """Ekstrak data dari format key-value pairs"""
        try:
            extracted_rows = []
            current_record = {}
            
            print(f"🔍 Processing {len(df)} rows for key-value extraction...")
            
            for row_idx in range(len(df)):
                row_data = df.iloc[row_idx]
                
                # Cari key-value pairs dalam baris ini
                # Format: key di kolom 0, value di kolom 1, atau key di kolom 2, value di kolom 3
                for col_idx in range(0, min(len(row_data), 4), 2):
                    if col_idx + 1 < len(row_data):
                        key_cell = row_data.iloc[col_idx]
                        value_cell = row_data.iloc[col_idx + 1]
                        
                        if pd.notna(key_cell) and pd.notna(value_cell):
                            key = str(key_cell).strip()
                            value = str(value_cell).strip()
                            
                            # Skip jika value kosong atau hanya ":"
                            if value and value != ':':
                                # Bersihkan value dari ":" di depan
                                if value.startswith(': '):
                                    value = value[2:].strip()
                                elif value.startswith(':'):
                                    value = value[1:].strip()
                                
                                # Map key ke field yang dikenal
                                mapped_field = self._map_key_to_field(key)
                                if mapped_field:
                                    current_record[mapped_field] = value
                                    print(f"  📝 Found {mapped_field}: {value}")
                
                # Jika baris ini berisi data transaksi (ada jumlah dan nilai)
                if self._is_transaction_row(row_data):
                    print(f"  💰 Transaction row detected at row {row_idx}")
                    # Tambahkan record yang sudah dikumpulkan
                    if current_record:
                        # Tambahkan data transaksi dari baris ini
                        transaction_data = self._extract_transaction_data(row_data)
                        current_record.update(transaction_data)
                        
                        extracted_rows.append(current_record.copy())
                        print(f"  ✅ Added record: {current_record}")
                        current_record = {}  # Reset untuk record berikutnya
                    else:
                        # Jika tidak ada current_record, buat record baru dengan data transaksi saja
                        print(f"  ⚠️ No current record, creating new one from transaction data")
                        transaction_data = self._extract_transaction_data(row_data)
                        if transaction_data:
                            # Buat record minimal dengan data yang tersedia
                            minimal_record = {
                                'jenis_biaya': transaction_data.get('jenis_biaya', ''),
                                'keterangan': transaction_data.get('keterangan', ''),
                                'jumlah': transaction_data.get('jumlah', ''),
                                'nilai': transaction_data.get('nilai', ''),
                                'sub_total': transaction_data.get('sub_total', '')
                            }
                            extracted_rows.append(minimal_record)
                            print(f"  ✅ Added minimal record: {minimal_record}")
                
                # Cek juga untuk baris yang berisi total/subtotal
                elif self._is_total_row(row_data):
                    print(f"  💰 Total row detected at row {row_idx}")
                    if current_record:
                        # Tambahkan data total dari baris ini
                        total_data = self._extract_total_data(row_data)
                        current_record.update(total_data)
                        
                        extracted_rows.append(current_record.copy())
                        print(f"  ✅ Added record with total: {current_record}")
                        current_record = {}  # Reset untuk record berikutnya
            
            # Tambahkan record terakhir jika ada
            if current_record:
                extracted_rows.append(current_record)
                print(f"  ✅ Added final record: {current_record}")
            
            print(f"📊 Extracted {len(extracted_rows)} records from key-value format")
            return extracted_rows
            
        except Exception as e:
            print(f"⚠️ Warning: Error extracting key-value data: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def _map_key_to_field(self, key):
        """Map key dari format key-value ke field yang dikenal"""
        key_lower = key.lower()
        
        # Mapping berdasarkan contoh data dari sampledata.xlsx
        if 'nomor tagihan' in key_lower:
            return 'nomor_tagihan'
        elif 'nomor registrasi' in key_lower:
            return 'nomor_registrasi'
        elif 'nama pasien' in key_lower:
            return 'nama_pasien'
        elif 'nama' in key_lower and 'pasien' in key_lower:
            return 'nama_pasien'
        elif 'pasien' in key_lower:
            return 'nama_pasien'
        elif 'tanggal registrasi' in key_lower:
            return 'tanggal_registrasi'
        elif 'kelas / kamar' in key_lower:
            return 'kelas_kamar'
        elif 'penjamin bayar' in key_lower:
            return 'penjamin_bayar'
        elif 'tanggal keluar' in key_lower:
            return 'tanggal_keluar'
        elif 'kelas dijamin' in key_lower:
            return 'kelas_dijamin'
        elif 'keterangan' in key_lower:
            return 'keterangan'
        elif 'jumlah' in key_lower:
            return 'jumlah'
        elif 'nilai' in key_lower:
            return 'nilai'
        elif 'biaya kamar' in key_lower:
            return 'jenis_biaya'
        elif 'room charge' in key_lower:
            return 'jenis_biaya'
        
        return None
    
    def _calculate_total_billed(self, data_row):
        """Hitung total billed dari tarif dikali quantity"""
        try:
            # Ambil nilai tarif dan quantity
            tarif = data_row.get('nilai', '')
            quantity = data_row.get('jumlah', '')
            
            # Cek apakah kedua nilai tersedia
            if not tarif or not quantity:
                return None
            
            # Bersihkan nilai tarif dari format currency
            tarif_clean = self._clean_currency_value(tarif)
            quantity_clean = self._clean_numeric_value(quantity)
            
            if tarif_clean is not None and quantity_clean is not None:
                # Hitung total
                total = tarif_clean * quantity_clean
                return f"{total:,.0f}"
            
            return None
            
        except Exception as e:
            print(f"⚠️ Warning: Error calculating total billed: {e}")
            return None
    
    def _clean_currency_value(self, value):
        """Bersihkan nilai currency untuk perhitungan"""
        try:
            if pd.isna(value) or value == '':
                return None
            
            # Konversi ke string
            value_str = str(value).strip()
            
            # Hapus karakter currency dan spasi
            value_str = value_str.replace('Rp', '').replace(' ', '').replace(',', '').replace('-', '').replace('.', '')
            
            # Cek apakah ada angka
            if re.search(r'\d', value_str):
                # Ekstrak angka saja
                numbers = re.findall(r'\d+', value_str)
                if numbers:
                    return float(numbers[0])
            
            return None
            
        except Exception as e:
            print(f"⚠️ Warning: Error cleaning currency value '{value}': {e}")
            return None
    
    def _clean_numeric_value(self, value):
        """Bersihkan nilai numeric untuk perhitungan"""
        try:
            if pd.isna(value) or value == '':
                return None
            
            # Konversi ke string
            value_str = str(value).strip()
            
            # Hapus karakter non-numeric kecuali titik dan minus
            value_str = re.sub(r'[^\d.-]', '', value_str)
            
            # Cek apakah ada angka
            if re.search(r'\d', value_str):
                return float(value_str)
            
            return None
            
        except Exception as e:
            print(f"⚠️ Warning: Error cleaning numeric value '{value}': {e}")
            return None
    
    def _extract_transaction_data(self, row_data):
        """Ekstrak data transaksi dari baris yang berisi jumlah dan nilai"""
        try:
            transaction_data = {}
            
            # Cek kolom 0: Jenis biaya
            if pd.notna(row_data.iloc[0]):
                transaction_data['jenis_biaya'] = self._clean_value(row_data.iloc[0])
            
            # Cek kolom 1: Keterangan
            if pd.notna(row_data.iloc[1]):
                transaction_data['keterangan'] = self._clean_value(row_data.iloc[1])
            
            # Cek kolom 4: Jumlah
            if len(row_data) > 4 and pd.notna(row_data.iloc[4]):
                transaction_data['jumlah'] = self._clean_value(row_data.iloc[4])
            
            # Cek kolom 5: Nilai
            if len(row_data) > 5 and pd.notna(row_data.iloc[5]):
                transaction_data['nilai'] = self._clean_value(row_data.iloc[5])
            
            # Cek kolom 7: Total (jika ada)
            if len(row_data) > 7 and pd.notna(row_data.iloc[7]):
                transaction_data['sub_total'] = self._clean_value(row_data.iloc[7])
            
            print(f"    💳 Transaction data: {transaction_data}")
            return transaction_data
            
        except Exception as e:
            print(f"⚠️ Warning: Error extracting transaction data: {e}")
            return {}
    
    def _is_transaction_row(self, row_data):
        """Cek apakah baris berisi data transaksi (ada jumlah dan nilai)"""
        try:
            # Cek apakah ada kolom jumlah dan nilai yang terisi
            if len(row_data) > 5:
                jumlah = row_data.iloc[4]  # Kolom 4: JUMLAH
                nilai = row_data.iloc[5]   # Kolom 5: NILAI
                
                # Hanya deteksi sebagai transaksi jika kedua kolom terisi dan bukan header
                if pd.notna(jumlah) and pd.notna(nilai):
                    jumlah_str = str(jumlah).strip()
                    nilai_str = str(nilai).strip()
                    
                    # Skip jika ini header (JUMLAH, NILAI)
                    if jumlah_str.upper() in ['JUMLAH', 'NILAI'] or nilai_str.upper() in ['JUMLAH', 'NILAI']:
                        return False
                    
                    # Cek apakah ada angka atau format currency
                    if re.search(r'\d', jumlah_str) and re.search(r'\d', nilai_str):
                        return True
            
            return False
        except Exception as e:
            print(f"⚠️ Warning: Error in _is_transaction_row: {e}")
            return False
    
    def _is_total_row(self, row_data):
        """Cek apakah baris berisi data total/subtotal"""
        try:
            if len(row_data) > 0:
                first_cell = str(row_data.iloc[0]).strip()
                # Cek apakah ini baris total atau subtotal
                if 'total' in first_cell.lower() or 'subtotal' in first_cell.lower():
                    return True
            return False
        except Exception as e:
            print(f"⚠️ Warning: Error in _is_total_row: {e}")
            return False
    
    def _extract_total_data(self, row_data):
        """Ekstrak data total dari baris total/subtotal"""
        try:
            total_data = {}
            
            # Cek kolom 0: Jenis biaya (Subtotal, BIAYA VISITE, dll)
            if pd.notna(row_data.iloc[0]):
                total_data['jenis_biaya'] = self._clean_value(row_data.iloc[0])
            
            # Cek kolom 7: Total nilai
            if len(row_data) > 7 and pd.notna(row_data.iloc[7]):
                total_data['sub_total'] = self._clean_value(row_data.iloc[7])
            
            print(f"    💳 Total data: {total_data}")
            return total_data
            
        except Exception as e:
            print(f"⚠️ Warning: Error extracting total data: {e}")
            return {}
    
    def _extract_raw_data(self, df, sheet_name):
        """Ekstrak data dari DataFrame tanpa header yang jelas"""
        try:
            extracted_rows = []
            
            # Coba ekstrak data berdasarkan posisi kolom
            for row_idx in range(len(df)):
                row_data = df.iloc[row_idx]
                
                # Skip baris yang kosong
                if row_data.isna().all():
                    continue
                
                extracted_row = {}
                
                # Coba ekstrak berdasarkan posisi kolom
                if len(row_data) >= 2:
                    # Kolom 0: Jenis biaya
                    if pd.notna(row_data.iloc[0]):
                        extracted_row['jenis_biaya'] = self._clean_value(row_data.iloc[0])
                    
                    # Kolom 1: Keterangan
                    if pd.notna(row_data.iloc[1]):
                        extracted_row['keterangan'] = self._clean_value(row_data.iloc[1])
                    
                    # Kolom 4: Jumlah
                    if len(row_data) > 4 and pd.notna(row_data.iloc[4]):
                        extracted_row['jumlah'] = self._clean_value(row_data.iloc[4])
                    
                    # Kolom 5: Nilai
                    if len(row_data) > 5 and pd.notna(row_data.iloc[5]):
                        extracted_row['nilai'] = self._clean_value(row_data.iloc[5])
                    
                    # Kolom 7: Total
                    if len(row_data) > 7 and pd.notna(row_data.iloc[7]):
                        extracted_row['sub_total'] = self._clean_value(row_data.iloc[7])
                
                if extracted_row:  # Hanya tambahkan jika ada data
                    extracted_rows.append(extracted_row)
            
            print(f"📊 Extracted {len(extracted_rows)} records from raw data")
            return extracted_rows
            
        except Exception as e:
            print(f"⚠️ Warning: Error extracting raw data: {e}")
            return []
    
    def _transform_to_output_format(self, extracted_data, analysis):
        """Transform data yang diekstrak ke format output yang diinginkan"""
        try:
            output_rows = []
            
            print(f"🔄 Transforming {len(extracted_data)} extracted records to output format")
            
            for idx, data_row in enumerate(extracted_data):
                output_row = {}
                
                # Map field yang diekstrak ke kolom output
                for output_col in self.output_columns:
                    mapped_value = self._map_extracted_field_to_output(output_col, data_row, analysis)
                    output_row[output_col] = mapped_value
                
                output_rows.append(output_row)
                
                # Debug: print first few rows
                if idx < 3:
                    print(f"  📋 Row {idx}: {list(output_row.values())[:5]}...")
            
            output_df = pd.DataFrame(output_rows)
            print(f"✅ Transformed to DataFrame with shape: {output_df.shape}")
            
            # Apply forward fill untuk kolom-kolom yang diminta
            output_df = self._apply_forward_fill(output_df)
            
            return output_df
            
        except Exception as e:
            print(f"❌ Error in transform_to_output_format: {e}")
            import traceback
            traceback.print_exc()
            # Return empty DataFrame as fallback
            return pd.DataFrame(columns=self.output_columns)
    
    def _map_extracted_field_to_output(self, output_col, data_row, analysis):
        """Map field yang diekstrak ke kolom output"""
        try:
            # Mapping berdasarkan field yang diekstrak
            field_mapping = {
                'CLIENT NAME': 'nama_pasien',
                'CLIENTS INVOICE NUMBER': 'nomor_tagihan',
                'CLIENTSREGISTER NUMBER': 'nomor_registrasi',
                'admission': 'tanggal_registrasi',
                'discharge': 'tanggal_keluar',
                'KELAS': 'kelas_kamar',
                'TARIFF': 'nilai',
                'QUANTITY': 'jumlah',
                'TOTAL BILLED': 'sub_total',
                'SERVICECODE DESCRIPTION': 'keterangan',
                'GIVEN DATE (month, day, year)': 'tanggal'
            }
            
            if output_col in field_mapping:
                source_field = field_mapping[output_col]
                if source_field in data_row:
                    value = data_row[source_field]
                    if pd.notna(value) and value != '':
                        return value
            
            # Mapping tambahan untuk field yang spesifik
            if output_col == 'CLIENT NAME' and 'nama_pasien' in data_row:
                return data_row['nama_pasien']
            elif output_col == 'CLIENTS INVOICE NUMBER' and 'nomor_tagihan' in data_row:
                return data_row['nomor_tagihan']
            elif output_col == 'CLIENTSREGISTER NUMBER' and 'nomor_registrasi' in data_row:
                return data_row['nomor_registrasi']
            elif output_col == 'KELAS' and 'kelas_kamar' in data_row:
                return data_row['kelas_kamar']
            elif output_col == 'TARIFF' and 'nilai' in data_row:
                return data_row['nilai']
            elif output_col == 'QUANTITY' and 'jumlah' in data_row:
                return data_row['jumlah']
            elif output_col == 'TOTAL BILLED':
                # Hitung total billed dari tarif dikali quantity
                calculated_total = self._calculate_total_billed(data_row)
                if calculated_total:
                    return calculated_total
                # Fallback ke sub_total jika ada
                elif 'sub_total' in data_row:
                    return data_row['sub_total']
            elif output_col == 'SERVICECODE DESCRIPTION':
                # Apply service code classification logic
                return self._classify_service_code(data_row)
            elif output_col == 'TARIFF DESCRIPTION':
                # Return empty for tariff description
                return ''
            elif output_col == 'SERVICECODE':
                # Apply service code classification logic for service code
                return self._classify_service_code_value(data_row)
            
            # Default values
            return self._get_default_value(output_col, data_row, None)
            
        except Exception as e:
            print(f"⚠️ Warning: Error mapping field {output_col}: {e}")
            return self._get_default_value(output_col, data_row, None)
    
    def _classify_service_code(self, data_row):
        """Klasifikasi service code description berdasarkan jenis_biaya"""
        try:
            # Ambil jenis_biaya dari data row
            jenis_biaya = data_row.get('jenis_biaya', '')
            keterangan = data_row.get('keterangan', '')
            
            # Gabungkan jenis_biaya dan keterangan untuk analisis
            combined_text = f"{jenis_biaya} {keterangan}".lower()
            
            # Keywords untuk Alkes/Peralatan
            alkes_keywords = [
                'peralatan', 'alkes', 'alat', 'equipment', 'medical device',
                'medical equipment', 'device', 'instrumen', 'instrument',
                'pump', 'syringe', 'infus', 'oksigen', 'oxygen', 'catheter',
                'canul', 'tubee', 'extension', 'threeway', 'combopack',
                'spuit', 'syringe', 'kertas usg', 'pd gel', 'kasa'
            ]
            
            # Keywords untuk Obat
            obat_keywords = [
                'obat', 'medicine', 'drug', 'medication', 'farmasi', 'pharmacy',
                'tablet', 'kapsul', 'sirup', 'injeksi', 'injection', 'tab',
                'mg', 'ml', 'cc', 'nifedipin', 'candesartan', 'furosemide',
                'isosorbide', 'betadine', 'alcohol', 'aquabidest', 'new diatabs'
            ]
            
            # Cek apakah mengandung keywords Alkes
            if any(keyword in combined_text for keyword in alkes_keywords):
                return 'Alkes'
            
            # Cek apakah mengandung keywords Obat
            elif any(keyword in combined_text for keyword in obat_keywords):
                return 'Obat'
            
            # Jika tidak ada keywords yang cocok, return kosong
            else:
                return ''
                
        except Exception as e:
            print(f"⚠️ Warning: Error in service code classification: {e}")
            return ''
    
    def _classify_service_code_value(self, data_row):
        """Klasifikasi service code value berdasarkan jenis_biaya"""
        try:
            # Ambil jenis_biaya dari data row
            jenis_biaya = data_row.get('jenis_biaya', '')
            keterangan = data_row.get('keterangan', '')
            
            # Gabungkan jenis_biaya dan keterangan untuk analisis
            combined_text = f"{jenis_biaya} {keterangan}".lower()
            
            # Keywords untuk Alkes/Peralatan
            alkes_keywords = [
                'peralatan', 'alkes', 'alat', 'equipment', 'medical device',
                'medical equipment', 'device', 'instrumen', 'instrument',
                'pump', 'syringe', 'infus', 'oksigen', 'oxygen', 'catheter',
                'canul', 'tubee', 'extension', 'threeway', 'combopack',
                'spuit', 'syringe', 'kertas usg', 'pd gel', 'kasa'
            ]
            
            # Keywords untuk Obat
            obat_keywords = [
                'obat', 'medicine', 'drug', 'medication', 'farmasi', 'pharmacy',
                'tablet', 'kapsul', 'sirup', 'injeksi', 'injection', 'tab',
                'mg', 'ml', 'cc', 'nifedipin', 'candesartan', 'furosemide',
                'isosorbide', 'betadine', 'alcohol', 'aquabidest', 'new diatabs'
            ]
            
            # Cek apakah mengandung keywords Alkes
            if any(keyword in combined_text for keyword in alkes_keywords):
                return 'Alkes'
            
            # Cek apakah mengandung keywords Obat
            elif any(keyword in combined_text for keyword in obat_keywords):
                return 'Obat'
            
            # Jika tidak ada keywords yang cocok, return kosong
            else:
                return ''
                
        except Exception as e:
            print(f"⚠️ Warning: Error in service code value classification: {e}")
            return ''
    
    def _clean_value(self, value):
        """Membersihkan dan memformat nilai data"""
        if pd.isna(value) or value == '':
            return ''
        
        # Handle infinity dan NaN
        if isinstance(value, float):
            if pd.isinf(value) or pd.isna(value):
                return ''
            if value.is_integer():
                return str(int(value))
            return str(value)
        
        # Konversi ke string dan bersihkan
        if isinstance(value, (int, float)):
            return str(value)
        
        # Bersihkan string
        if isinstance(value, str):
            # Bersihkan whitespace dan karakter tidak valid
            cleaned = value.strip()
            cleaned = cleaned.replace('\x00', '').replace('\n', ' ').replace('\r', '')
            
            # Handle format currency
            if 'Rp' in cleaned:
                # Ekstrak angka dari format "Rp 75,000"
                numbers = re.findall(r'[\d,]+', cleaned)
                if numbers:
                    return numbers[0]
            
            return cleaned
        
        # Konversi ke string dan bersihkan
        try:
            result = str(value)
            return result.replace('\x00', '').replace('\n', ' ').replace('\r', '')
        except:
            return ''
    
    def _apply_forward_fill(self, df):
        """Apply forward fill untuk kolom-kolom yang diminta"""
        try:
            print(f"🔄 Applying forward fill to specified columns...")
            
            # Kolom-kolom yang akan di-forward fill
            forward_fill_columns = [
                'CLIENT NAME',
                'CLIENTS INVOICE NUMBER', 
                'CLIENTSREGISTER NUMBER',
                'KELAS'
            ]
            
            # Cek kolom mana yang ada dalam DataFrame
            available_columns = [col for col in forward_fill_columns if col in df.columns]
            
            if not available_columns:
                print(f"⚠️ No forward fill columns found in DataFrame")
                return df
            
            print(f"📊 Forward filling columns: {available_columns}")
            
            # Buat copy DataFrame untuk modifikasi
            df_filled = df.copy()
            
            # Apply forward fill untuk setiap kolom
            for col in available_columns:
                print(f"  🔄 Forward filling column: {col}")
                
                # Cek apakah ada nilai yang perlu di-forward fill
                empty_mask = df_filled[col].isna() | (df_filled[col] == '')
                
                if empty_mask.any():
                    print(f"    📝 Found {empty_mask.sum()} empty cells in {col}")
                    
                    # Forward fill menggunakan pandas ffill()
                    df_filled[col] = df_filled[col].replace('', pd.NA).ffill()
                    
                    # Cek hasil forward fill
                    filled_count = (df_filled[col].notna() & (df_filled[col] != '')).sum()
                    print(f"    ✅ After forward fill: {filled_count} non-empty cells")
                else:
                    print(f"    ✅ Column {col} already has all values filled")
            
            # Debug: tampilkan beberapa baris untuk verifikasi
            print(f"📋 Sample data after forward fill:")
            for idx, row in df_filled.head(5).iterrows():
                client_name = row.get('CLIENT NAME', 'N/A')
                invoice_number = row.get('CLIENTS INVOICE NUMBER', 'N/A')
                register_number = row.get('CLIENTSREGISTER NUMBER', 'N/A')
                print(f"    Row {idx}: CLIENT NAME='{client_name}', INVOICE='{invoice_number}', REGISTER='{register_number}'")
            
            return df_filled
            
        except Exception as e:
            print(f"⚠️ Warning: Error in forward fill: {e}")
            import traceback
            traceback.print_exc()
            return df
    
    def _get_default_value(self, column_name, row, df):
        """Memberikan nilai default untuk kolom yang tidak ada di input"""
        if column_name == 'GIVEN DATE (month, day, year)':
            return datetime.now().strftime('%m/%d/%Y')
        # elif column_name == 'CLIENTS SEX':
        #     return 'L'  # Default value
        elif column_name == 'admission':
            return datetime.now().strftime('%m/%d/%Y')
        elif column_name == 'discharge':
            return datetime.now().strftime('%m/%d/%Y')
        elif column_name == 'LoS':
            return '0'  # Length of Stay default
        # elif column_name == 'PHYSICIAN DESCRIPTION (DPJP/IGD/POLICLINIC)':
        #     return 'IGD'  # Default berdasarkan contoh data
        else:
            return ''
    
    def _create_output_file(self, df, input_filepath):
        """Membuat file Excel output"""
        try:
            # Buat nama file output
            input_filename = os.path.basename(input_filepath)
            name_without_ext = os.path.splitext(input_filename)[0]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"processed_{name_without_ext}_{timestamp}.xlsx"
            
            # Buat direktori output jika belum ada
            output_dir = 'outputs'
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            output_filepath = os.path.join(output_dir, output_filename)
            
            # Tulis ke file Excel
            with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Processed Data', index=False)
                
                # Auto-adjust column widths dan styling header
                worksheet = writer.sheets['Processed Data']
                
                # Import openpyxl styles
                from openpyxl.styles import PatternFill, Font
                
                # Define colors
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                black_font = Font(color='000000')
                
                # Columns that should have red background
                red_columns = ['PROVID', 'PROVIDER_NAME', 'HELPER']
                
                # Apply styling to header row (row 1)
                for col_idx, column in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column_letter = column[0].column_letter
                    header_cell = worksheet.cell(row=1, column=col_idx)
                    
                    # Check if this column should be red
                    if header_cell.value in red_columns:
                        header_cell.fill = red_fill
                        header_cell.font = black_font
                    else:
                        header_cell.fill = yellow_fill
                        header_cell.font = black_font
                    
                    # Auto-adjust column width
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return output_filepath
            
        except Exception as e:
            raise Exception(f"Error membuat file output: {str(e)}")
    
    def _clean_dataframe(self, df):
        """Membersihkan DataFrame dari karakter tidak valid dan data yang bermasalah"""
        try:
            # Bersihkan nama kolom
            df.columns = [str(col).replace('\x00', '').replace('\n', ' ').replace('\r', '').strip() for col in df.columns]
            
            # Bersihkan data dalam setiap kolom
            for col in df.columns:
                if df[col].dtype == 'object':
                    # Konversi ke string dan bersihkan
                    df[col] = df[col].astype(str).apply(lambda x: 
                        x.replace('\x00', '').replace('\n', ' ').replace('\r', '') if pd.notna(x) else ''
                    )
                elif df[col].dtype in ['int64', 'float64']:
                    # Handle NaN dan infinity values
                    df[col] = df[col].replace([np.inf, -np.inf], np.nan)
                    df[col] = df[col].fillna('')
            
            return df
        except Exception as e:
            print(f"Warning: Error saat membersihkan DataFrame: {e}")
            return df
    
    def analyze_data_structure(self, filepath):
        """Menganalisis struktur data Excel untuk pembelajaran"""
        try:
            df = pd.read_excel(filepath, engine='openpyxl')
            
            analysis = {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'column_info': {},
                'data_patterns': {},
                'sample_values': {}
            }
            
            for col in df.columns:
                col_data = df[col].dropna()
                
                # Informasi kolom
                analysis['column_info'][col] = {
                    'data_type': str(df[col].dtype),
                    'non_null_count': len(col_data),
                    'null_count': df[col].isna().sum(),
                    'unique_values': col_data.nunique()
                }
                
                # Sample values
                if len(col_data) > 0:
                    analysis['sample_values'][col] = col_data.head(3).tolist()
                
                # Pattern analysis
                if len(col_data) > 0:
                    if df[col].dtype in ['int64', 'float64']:
                        analysis['data_patterns'][col] = 'numeric'
                    elif col_data.str.contains('Rp', na=False).any():
                        analysis['data_patterns'][col] = 'currency'
                    elif col_data.str.contains(r'\d{2}/\d{2}/\d{4}', na=False).any():
                        analysis['data_patterns'][col] = 'date'
                    else:
                        analysis['data_patterns'][col] = 'text'
            
            return analysis
            
        except Exception as e:
            raise Exception(f"Error menganalisis struktur data: {str(e)}")
